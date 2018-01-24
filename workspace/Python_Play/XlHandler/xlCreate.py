'''
Created on Oct 22, 2015

@author: marshall
'''
from openpyxl import Workbook
from openpyxl import load_workbook


class WbCreate(object):
    '''
    This creates a XL workbook using openpyxl
    '''

    wb = Workbook()

    def __init__(self):
        
        ## Create the workbook
        
        
        
        self.wb = self.loadtemplate('/Users/marshall/workspace/Python_Play/XlHandler/MAQI_template.xlsx')
        ## work book is always created with 1 active sheet. you can add more with wb.create_sheet()
        ws = self.wb.active  ## must grab the active sheet. 
        
        ## ws.title = "Sheet 1" ## set the title of the work sheet
        ## ws2 = self.wb.create_sheet("Sheet 2") ## create sheet 2
        ## ws2.sheet_properties.tabColor = "1072BA"  ## change the tab color
        ## ws['A1'] = 42
        ## ws.append([1,2,3])
        self.wb.save(filename='/Users/marshall/workspace/Python_Play/XlHandler/New_MAQI.xlsx')
        
    def saveWb(self):
        self.wb.save(filename='/Users/marshall/workspace/Python_Play/XlHandler/New_MAQI.xlsx')
         
    def loadtemplate(self, filepath):
        
        ### load a template excel file
         
        # Open the file from command line
        try:
            wb = load_workbook(filepath)
            print "Opening File: " + filepath + "\n"
        except IOError:
            print "ERROR: Cannot open file at path: " + filepath
            return wb
    
    def createtemplate(self):
        wb = Workbook()
        templatesheet = wb.active()
        templatesheet.title = "Template"
        summarysheet = wb.create_shee("Summary")
        summarysheet.sheet_properties.tabColor = "1072BA"  ## change the tab color
        
        #create the template sheet. 
        #Here is the header
        templatesheet['A1'] = "index"
        templatesheet['B1'] = "Group"
        templatesheet['C1'] = "Item"
        templatesheet['D1'] = "Weighting"
        templatesheet['E1'] = "Percentage"
        templatesheet['F1'] = "Score"
        templatesheet['G1'] = "Applicable?"
        templatesheet['H1'] = "Notes"
        
class WbUtils(object):
    ''' used for utilites to work on work book
    '''
    def __init__(self, params):
        '''
        Constructor
        '''
        
    def wbtodb(self, wb):
        #Load a work book into a db
        #wb = Workbook(object)
        return 
        
    def sheetsearch(self, txt):
        '''
        searches for text in a sheet. may need to do something different for formula?
        '''
        " put in for loop for searching a sheet. "
        cell = 'A1'
        
        return cell
    
    # this takes in a cell and returns a cell one column higher
    def incrementColumn(self, cell):
        col = chr(ord(list(cell)[0]) +1)
        row = list(cell[1])
        return (col + row)
    # this takes in a cell and returns a cell one column lower
    def decrementColumn(self, cell):
        col = chr(ord(list(cell)[0]) - 1)
        row = list(cell[1])
        return (col + row)
    # this takes in a cell and returns a cell one column higher
    def incrementRow(self, cell):
        col = chr(ord(list(cell)[0]) +1)
        row = list(cell[1])
        return (col + row)
    # this takes in a cell and returns a cell one column lower
    def decrementRow(self, cell):
        col = chr(ord(list(cell)[0]) - 1)
        row = list(cell[1])
        return (col + row)
    
    
        