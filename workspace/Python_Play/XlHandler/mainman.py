#!/usr/local/bin/python2.7
# encoding: utf-8
'''
HelloWorld.helloworld -- This is just a short test of using python in eclipse
It defines classes_and_methods

@author:     George Marshall

@copyright:  2015. All rights reserved.

@contact:    george.marshall@here.com
@deffield    updated: Updated
'''

import sys
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import get_column_letter

def createSummaryPage(wb):
    ## this is to create the summary page for all workbooks.
    summarySheet = wb.create_sheet("Summary", 1)
    for sheet in wb: 
        summarySheet.append((sheet.title) + ":" + str(sheet.cell('D27').value)) # careful about mixing strings with cell values

def consolidateWorkbooks(self): 
        ## this is to  bring all the workbooks together into one. 
        print "test" 
        
def createTemplateSheet(self):
    ## this is to create a template sheet. Should read from file to create fields.
    print self
    
def writeToDb(self):
    ## write spread sheets to a db for analysis. 
    print self
   
def main(argv=None): # IGNORE:C0111
    try:
        ## Open the workbook template load from a file
        # wbtemplate = load_workbook('./MAQI_template.xlsx')
            newwb = load_workbook('./new_maqi.xlsx', data_only=True)  # must have data_only=True if you want to see cell value and not formula
        
        #debug
        ## print "Here are the sheet names:" + str(newwb.get_sheet_names())
        ## sheetnames = str(newwb.get_sheet_names())
        
        ## modify the work book template.    
        ## ws2 = wbtemplate.create_sheet(title="Leaf")
        ## Delete?  sheet_ranges = newwb['AET']
        ## Delete ?? print(sheet_ranges['B2'].value) + "\n"
        
        ## ws = newwb.get_sheet_by_name(newwb.get_sheet_names()[3])
        ## print "sheet name:" + newwb.get_sheet_names()[3]
        ## to print a cell value print ws['D27'].value  
        
        # print every sheet in a cell  creates a key value pair. DICT
        ## for sheet in newwb: 
        ##    print (sheet.title) + ":" + str(sheet.cell('D27').value) # careful about mixing strings with cell values
            createSummaryPage(newwb)
            newwb.save('./Modified.xlsx')
    except KeyboardInterrupt:
        ### handle keyboard interrupt ###
        return 0
    except Exception, e:
        sys.stderr.write(repr(e) + "\n")
        return 2
    
 
#TODO Load the work book sheets for the directory in the arguments. 
            
if __name__ == "__main__":
    sys.exit(main())
    
